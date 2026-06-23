import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR = os.path.dirname(os.path.abspath(__file__))

def load(name):
    with open(os.path.join(DIR, name), encoding="utf-8-sig") as f:
        return json.load(f)

full = load("players_full.json")          # list of detailed player objects
rnd = load("players_round1.json")["items"]  # priceChange / trend / pointsChange

# map playerId -> round info
rmap = {r["playerId"]: r for r in rnd}

POS_ORDER = {"Keeper": 0, "Målmand": 0, "Forsvar": 1, "Midtbane": 2, "Angreb": 3}

rows = []
for p in full:
    pid = p["id"]
    person = p.get("person") or {}
    team = p.get("team") or {}
    pos = p.get("position") or {}
    name = (str(person.get("firstName") or "").strip() + " " +
            str(person.get("lastName") or "").strip()).strip()
    price = p.get("price") or 0
    start = p.get("startPrice") or 0
    growth = price - start
    r = rmap.get(pid, {})
    rows.append({
        "Navn": name,
        "Hold": team.get("name") or "",
        "Position": pos.get("title") or pos.get("name") or "",
        "Pris": price,
        "Startpris": start,
        "Vaerdivaekst": growth,
        "Vaekst_pct": (growth / start) if start else 0,
        "Vaekst_seneste_runde": r.get("priceChange") or 0,
        "Trend": r.get("trend") or 0,
        "Point": p.get("points") or 0,
        "Popularitet_pct": (p.get("popularity") or 0),
        "Skadet": "Ja" if p.get("isInjured") else "",
        "Ude_af_spil": "Ja" if (p.get("isOutOfGame") or p.get("isEliminated")) else "",
    })

# sort: position, then price desc
rows.sort(key=lambda x: (POS_ORDER.get(x["Position"], 9), -x["Pris"]))

headers = [
    ("Navn", 26), ("Hold", 18), ("Position", 12),
    ("Pris", 14), ("Startpris", 14), ("Værdivækst", 14), ("Vækst %", 11),
    ("Vækst (seneste runde)", 18), ("Trend", 12),
    ("Point", 9), ("Popularitet %", 13), ("Skadet", 8), ("Ude af spil", 11),
]
keys = ["Navn","Hold","Position","Pris","Startpris","Vaerdivaekst","Vaekst_pct",
        "Vaekst_seneste_runde","Trend","Point","Popularitet_pct","Skadet","Ude_af_spil"]

wb = Workbook()
ws = wb.active
ws.title = "Spillere"

hdr_fill = PatternFill("solid", fgColor="1F2A44")
hdr_font = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, (label, width) in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c, value=label)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(c)].width = width

for i, row in enumerate(rows, start=2):
    for c, k in enumerate(keys, start=1):
        cell = ws.cell(row=i, column=c, value=row[k])
        cell.border = border
        if k in ("Pris","Startpris","Vaerdivaekst","Vaekst_seneste_runde","Trend"):
            cell.number_format = '#,##0'
        elif k == "Vaekst_pct":
            cell.number_format = '0.0%'
        elif k == "Popularitet_pct":
            cell.number_format = '0.00%'

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

# Summary sheet by team
from collections import defaultdict
agg = defaultdict(lambda: {"n":0,"sum":0})
for r in rows:
    agg[r["Hold"]]["n"] += 1
    agg[r["Hold"]]["sum"] += r["Pris"]
ws2 = wb.create_sheet("Hold-oversigt")
ws2.append(["Hold","Antal spillere","Samlet pris","Gns. pris"])
for c in range(1,5):
    cell = ws2.cell(row=1, column=c); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
for team in sorted(agg):
    a = agg[team]
    ws2.append([team, a["n"], a["sum"], round(a["sum"]/a["n"]) if a["n"] else 0])
for col, w in zip("ABCD",[20,14,16,14]):
    ws2.column_dimensions[col].width = w
for r in range(2, ws2.max_row+1):
    ws2.cell(row=r,column=3).number_format = '#,##0'
    ws2.cell(row=r,column=4).number_format = '#,##0'

out = os.path.join(DIR, "..", "Holdet_VM2026_spillere.xlsx")
out = os.path.abspath(out)
wb.save(out)
print("rows:", len(rows))
print("saved:", out)
